# import modules to interact with the os, manipulate json, work with time, produce specific audio alert, and make online requests
import os, json, time, simpleaudio, requests

# import pandas to work with dataframes
import pandas as pd

# import modules from openpyxl to work with excel
from openpyxl import Workbook, load_workbook

# import tqdm for a progress bar
from tqdm import tqdm

# import web3 to interact with EVM
from web3 import Web3

# import colorama to add color to printed texts
from colorama import init, Fore


# function to load config data
def load_config():
    # open the config.json file
    with open("./config.json", "r") as file:
        # parse json string and load it to a dictionary variable
        config = json.loads(file.read())
    return config


# get ABI based on a contract address and API
def get_abi(contract_address, port_addy):
    # compose API endpoint based on specific API endpoint format
    API_ENDPOINT = (
        port_addy + "?module=contract&action=getabi&address=" + str(contract_address)
    )
    # make a request for the ABI in JSON
    r = requests.get(url=API_ENDPOINT)
    response = r.json()
    # parse json string and load it to a dictionary variable
    abi = json.loads(response["result"])
    # wait a little so that the connection doesn't get blocked for spamming
    time.sleep(5)
    return abi


# function to get and store new ABIs
def store_new_abi(file_name, blockchain):
    # check to see if the ABI exists based on the filename if the ABI doesn't exist then get and save one
    if os.path.exists("./ABIs/" + str(file_name) + ".json") == False:
        # load config data
        config = load_config()
        # split the file name so that it can be used in subsequent lines
        split_fn = file_name.split("_")
        try:
            # check for a contract address
            CONTRACT_ADDRESS = config[blockchain][split_fn[0]][file_name]
        except KeyError:
            # if there isn't a contract address then assume it's a pair contract being targeted find it and get the contract address
            # split_fn = file_name.split("_")
            factory_name = "_".join([split_fn[0], split_fn[1]])
            # get the address from the config data
            factory_address = config[blockchain][split_fn[0]][factory_name]
            # load abi json
            with open("./ABIs/" + str(factory_name) + ".json", "r") as file:
                factory_abi = json.loads(file.read())
            # get the contract object
            factory_contract = getContract(blockchain, factory_address, factory_abi)
            # use contract object to get contract address for a pair, which can be used to obtain a sample ABI for all pairs created by that factory
            CONTRACT_ADDRESS = factory_contract.functions.allPairs(0).call()

        # use get_abi to get a sample abi
        abi = get_abi(
            contract_address=CONTRACT_ADDRESS,
            port_addy=config[blockchain]["abi_api"],
        )

        # dump result to a json file
        with open("./ABIs/" + str(file_name) + ".json", "w") as file:
            json.dump(abi, file)


# function to get the gas fees and gas limit
def check_gas_fee(blockchain):
    # get config data
    config = load_config()
    # use Web3 to connect to blockchain
    w3 = Web3(Web3.HTTPProvider(config[blockchain]["network"]["mainnet"]))
    # get gas price in wei
    gas_fee_wei = w3.eth.gas_price
    # get priority fee or miner tip in wei
    miner_tip_wei = w3.eth.max_priority_fee
    # get total fee in wei
    tip_and_gas_wei = gas_fee_wei + miner_tip_wei
    # get latest block
    latest_block = w3.eth.getBlock("latest")
    # define gas limit
    gas_limit = int(
        latest_block.gasLimit
        / (1 if len(latest_block.transactions) == 0 else len(latest_block.transactions))
    )
    # get the max total gas cost in wei, gwei, and ether
    max_total_gas_wei = tip_and_gas_wei * gas_limit
    max_total_gas_gwei = w3.fromWei(max_total_gas_wei, "gwei")
    max_total_gas_eth = w3.fromWei(max_total_gas_wei, "ether")

    print("")
    print(f"Base Gas = {gas_fee_wei} Wei")
    print(f"Miner's Tip = {miner_tip_wei} Wei")
    print(f"Gas fee = {tip_and_gas_wei} Wei")
    print(f"Gas limit = {gas_limit}")
    print(f"Gas Total = {max_total_gas_wei} Wei")
    print(f"Gas Total = {max_total_gas_gwei} Gwei")
    print(f"Gas Total = {max_total_gas_eth} Eth")
    print("")

    return tip_and_gas_wei, gas_limit, max_total_gas_eth


# builds a new contract based on abi and address
def getContract(blockchain, address, abi):
    # load config data
    config = load_config()
    # use Web3 to connect to blockchain
    w3 = Web3(Web3.HTTPProvider(config[blockchain]["network"]["mainnet"]))
    # make sure address is in acceptable
    address = w3.toChecksumAddress(address)
    # build contract object based on address and abi
    contract = w3.eth.contract(abi=abi, address=address)
    return contract


# prepare export dictionaries for function that scans blockchain for pair data
def prep_export_dict(pair_names, xch_names):
    # initialise dictionaries
    over_dict = {}
    skip_pair = {}

    # loop through the pair names
    for i in pair_names:
        # split the names into two parts
        split_pair_name = i.split("_")
        # initialise a temporary dictionary
        scanned_dict = {}
        # loop thorough target exchanges to create new fields each token and the purchasing price with reference to the base token of the chain
        for j in xch_names:
            scanned_dict[j + str("_") + str(split_pair_name[0])] = []
            scanned_dict[j + str("_") + str(split_pair_name[1])] = []
            scanned_dict[j + str("_buy_with_base")] = []

        scanned_dict["gross_perc_profit"] = []
        scanned_dict["potential_trade"] = []
        # note that for this to work in the rest of this application the trade path should start and end with base_token
        scanned_dict["trade_path"] = []
        over_dict[i] = scanned_dict
        skip_pair[i] = 0

    return over_dict, skip_pair


# function to estimate potential arbitrage percentage opportunity
def arb_value(over_dict, xch_names, pair, small_cap):
    cap_key0 = str(xch_names[0]) + str("_buy_with_base")
    cap_key1 = str(xch_names[1]) + str("_buy_with_base")
    gross_perc_profit = (
        (over_dict[pair][cap_key0][-1] - over_dict[pair][cap_key1][-1])
        / over_dict[pair][cap_key0][-1]
    ) * 100

    deductable = 0.8

    if gross_perc_profit < 0:
        gross_perc_profit = gross_perc_profit + deductable
        if gross_perc_profit > 0:
            gross_perc_profit = 0
    elif gross_perc_profit > 0:
        gross_perc_profit = gross_perc_profit - deductable
        if gross_perc_profit < 0:
            gross_perc_profit = 0

    if abs(gross_perc_profit) <= 2:
        potential_trade = False
        gross_perc_profit = 0
    else:
        potential_trade = True
        gross_perc_profit = gross_perc_profit

    if small_cap == True:
        potential_trade = False

    over_dict[pair]["gross_perc_profit"].append(gross_perc_profit)
    over_dict[pair]["potential_trade"].append(potential_trade)

    return over_dict


# get data on the price ratio and pool size of a specified pair from a specified exchange
def get_specific_pair(
    xch_name,
    blockchain,
    address,
    pair_name,
    base_token,
):
    # load pool sample abi json
    with open("./ABIs/" + str(xch_name) + "factory_pool.json", "r") as file:
        pool_abi = json.loads(file.read())

    with open("./ABIs/" + str(xch_name) + "_router.json", "r") as file:
        router_abi = json.loads(file.read())

    # load config file and get contract for a given pair on a given exchange
    config = load_config()
    pair_contract = getContract(blockchain, address, pool_abi)

    # get router contract
    router_contract = getContract(
        blockchain,
        config[blockchain][xch_name][str(xch_name) + str("_router")],
        router_abi,
    )

    amount_in = Web3.toWei(1, "ether")

    # get on chain data
    reserve = pair_contract.functions.getReserves().call()
    split_pair_name = pair_name.split("_")
    t0_reserve = reserve[0] / (10 ** config[blockchain][split_pair_name[0]])
    t1_reserve = reserve[1] / (10 ** config[blockchain][split_pair_name[1]])

    # ensure that swap ratio is always given relative to base token
    if split_pair_name[0] == base_token:
        base_token_address = pair_contract.functions.token0().call()
        other_token_address = pair_contract.functions.token1().call()
        base_reserve = split_pair_name[0]

        address_path = [base_token_address, other_token_address]
        get_amount_out = router_contract.functions.getAmountsOut(
            amount_in, address_path
        ).call()
        swap_ratio = get_amount_out[1] / (10 ** config[blockchain][split_pair_name[0]])

    elif split_pair_name[1] == base_token:
        base_token_address = pair_contract.functions.token1().call()
        other_token_address = pair_contract.functions.token0().call()
        base_reserve = split_pair_name[1]

        address_path = [base_token_address, other_token_address]
        get_amount_out = router_contract.functions.getAmountsOut(
            amount_in, address_path
        ).call()
        swap_ratio = get_amount_out[1] / (10 ** config[blockchain][split_pair_name[1]])

    return (
        t0_reserve,
        t1_reserve,
        swap_ratio,
        split_pair_name,
        base_token_address,
        other_token_address,
        base_reserve,
    )


# get info on the tokens in the pair
def get_t0t1_decimals(token0_address, token1_address, w3, port_addy):
    token0_abi = get_abi(contract_address=token0_address, port_addy=port_addy)
    token0_contract = w3.eth.contract(abi=token0_abi, address=token0_address)
    try:
        token0_decimals = token0_contract.functions.decimals().call()
    except:
        token0_decimals = 18
    try:
        token0_symbol = token0_contract.functions.symbol().call()
    except:
        token0_symbol = None

    # check token1
    token1_abi = get_abi(contract_address=token1_address, port_addy=port_addy)
    token1_contract = w3.eth.contract(abi=token1_abi, address=token1_address)
    try:
        token1_decimals = token1_contract.functions.decimals().call()
    except:
        token1_decimals = 18
    try:
        token1_symbol = token1_contract.functions.symbol().call()
    except:
        token1_symbol = None

    return (token0_decimals, token1_decimals, token0_symbol, token1_symbol)


# function to scan the factory contract and get all pairs - useful for prospecting viable arb pairs
# note that it is only useful for pairs with base token e.g. wbnb, otherwise changes to code will be required
def get_pairs_from_factory(
    file_name, blockchain, secondary_dex, selected_ids, save_name, base_token
):
    print("")
    col_list = [
        "DEX_pool_no",
        "t0_address",
        "t1_address",
        "DEX_name",
        "pool_address",
        "t0_reserves",
        "t1_reserves",
        "amountOut",
        "s_DEX_name",
        "s_pool_address",
        "s_t0_reserves",
        "s_t1_reserves",
        "s_amountOut",
        "balance",
        "arb",
    ]
    return_list = []

    wb = Workbook()
    ws = wb.active
    ws.append(col_list)
    wb.save(save_name)

    # load factory abi json
    with open("./ABIs/" + str(file_name) + ".json", "r") as file:
        factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(file_name) + "_pool.json", "r") as file:
        pool_abi = json.loads(file.read())

    with open("./ABIs/" + str(file_name.split("_")[0]) + "_router.json", "r") as file:
        router_abi = json.loads(file.read())

    # load factory abi json
    with open("./ABIs/" + str(secondary_dex) + ".json", "r") as file:
        sdex_factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(secondary_dex) + "_pool.json", "r") as file:
        sdex_pool_abi = json.loads(file.read())

    with open(
        "./ABIs/" + str(secondary_dex.split("_")[0]) + "_router.json", "r"
    ) as file:
        sdex_router_abi = json.loads(file.read())

    config = load_config()
    w3 = Web3(Web3.HTTPProvider(config[blockchain]["network"]["mainnet"]))
    factory_address = w3.toChecksumAddress(
        config[blockchain][file_name.split("_")[0]][file_name]
    )
    factory_contract = w3.eth.contract(abi=factory_abi, address=factory_address)

    dex_router_contract = w3.eth.contract(
        abi=router_abi,
        address=config[blockchain][file_name.split("_")[0]][
            str(file_name.split("_")[0]) + "_router"
        ],
    )

    if selected_ids == None:
        record_length = factory_contract.functions.allPairsLength().call()
        sample_range = list(range(record_length))
    else:
        sample_range = selected_ids

    sdex_address = w3.toChecksumAddress(
        config[blockchain][secondary_dex.split("_")[0]][secondary_dex]
    )
    sdex_contract = w3.eth.contract(abi=sdex_factory_abi, address=sdex_address)

    s_dex_router_contract = w3.eth.contract(
        abi=sdex_router_abi,
        address=config[blockchain][secondary_dex.split("_")[0]][
            str(secondary_dex.split("_")[0]) + "_router"
        ],
    )

    for i in tqdm(sample_range, "Downloading: ", leave=False):
        pool_address = factory_contract.functions.allPairs(i).call()
        pool_address = w3.toChecksumAddress(pool_address)
        pool_contract = w3.eth.contract(abi=pool_abi, address=pool_address)

        # get pool data
        reserves = pool_contract.functions.getReserves().call()
        token0_address = pool_contract.functions.token0().call()
        token1_address = pool_contract.functions.token1().call()

        reserves[0] = reserves[0]
        reserves[1] = reserves[1]

        # see if the same pool exists in pancake swap - secondary_dex
        try:
            sdex_pool_address = sdex_contract.functions.getPair(
                token0_address, token1_address
            ).call()
            sdex_pool_address = w3.toChecksumAddress(sdex_pool_address)
            sdex_pool_contract = w3.eth.contract(
                abi=sdex_pool_abi, address=sdex_pool_address
            )

            # get secondary pool data
            s_reserves = sdex_pool_contract.functions.getReserves().call()
            s_reserves[0] = s_reserves[0]
            s_reserves[1] = s_reserves[1]

            # determine how many other tokens the base token will get
            base_token_in = 1
            base_token_in = Web3.toWei(base_token_in, "ether")
            for addy in [token0_address, token1_address]:
                if addy != base_token:
                    other_token = addy
                    break
            trade_path = [base_token, other_token]
            amountOut = dex_router_contract.functions.getAmountsOut(
                base_token_in, trade_path
            ).call()[1]
            s_amountOut = s_dex_router_contract.functions.getAmountsOut(
                base_token_in, trade_path
            ).call()[1]

            # find the better value
            if amountOut > s_amountOut:
                end_trade = s_dex_router_contract.functions.getAmountsOut(
                    amountOut, [other_token, base_token]
                ).call()[1]
            elif amountOut < s_amountOut:
                end_trade = dex_router_contract.functions.getAmountsOut(
                    s_amountOut, [other_token, base_token]
                ).call()[1]
            else:
                end_trade = s_dex_router_contract.functions.getAmountsOut(
                    amountOut, [other_token, base_token]
                ).call()[1]

            arb = (end_trade - base_token_in) / base_token_in
            arb = arb - 0.00166

            # if other_token == "0xacFC95585D80Ab62f67A14C566C1b7a49Fe91167":
            #     arb = arb - 0.02
            # else:
            #     arb = arb - 0.00166

            if arb > 0:
                return_list = [
                    i,
                    token0_address,
                    token1_address,
                    file_name.split("_")[0],
                    pool_address,
                    reserves[0],
                    reserves[1],
                    amountOut,
                    secondary_dex.split("_")[0],
                    sdex_pool_address,
                    s_reserves[0],
                    s_reserves[1],
                    s_amountOut,
                    end_trade,
                    arb,
                ]

                book = load_workbook(save_name)
                sheet = book.active
                sheet.append(return_list)
                book.save(save_name)

                for i in range(5):
                    play_obj = simpleaudio.WaveObject.from_wave_file(
                        "mixkit-basketball-buzzer-1647.wav"
                    ).play()
                    play_obj.wait_done()
                    time.sleep(1)

        except:
            pass


def scan_by_name(
    pair_names,
    xch_names,
    blockchain,
    base_token,
):

    print("")

    # define time intervals to prevent spamming
    nap = 300
    hour = 5
    trade_path = ""
    small_cap = False

    best_set = {
        "pair_name": "",
        "trade_value": 0,
        "trade_path": "",
        "xch0": "",
        "xch1": "",
        "base_token": "",
        "other_token": "",
    }

    # get dictionaries
    over_dict, skip_pair = prep_export_dict(pair_names, xch_names)

    for step in range(hour):
        # for step in tqdm(range(hour), "Downloading: ", leave=True):
        if step > 0:
            time.sleep(nap)

        # for each pair search all exchanges provided
        for i in tqdm(pair_names, "Scanning: ", leave=False):
            if skip_pair[i] == 0:
                for j in xch_names:
                    config = load_config()
                    pair_address = config[blockchain][j]["pool_pairs"][i]
                    (
                        t0_reserve,
                        t1_reserve,
                        swap_ratio,
                        split_pair_name,
                        base_token_address,
                        other_token_address,
                        base_reserve,
                    ) = get_specific_pair(
                        xch_name=j,
                        blockchain=blockchain,
                        address=pair_address,
                        pair_name=i,
                        base_token=base_token,
                    )

                    # save onchain data to dictionary for export to excel
                    over_dict[i][j + str("_") + str(split_pair_name[0])].append(
                        t0_reserve
                    )
                    over_dict[i][j + str("_") + str(split_pair_name[1])].append(
                        t1_reserve
                    )
                    over_dict[i][j + str("_buy_with_base")].append(swap_ratio)

                # get arbitrage value
                over_dict = arb_value(
                    over_dict=over_dict,
                    xch_names=xch_names,
                    pair=i,
                    small_cap=small_cap,
                )

                # give a recommendation of trade path
                if over_dict[i]["potential_trade"][-1] == True:
                    # buy at xch0 and sell at xch1
                    if over_dict[i]["gross_perc_profit"][-1] > 0:
                        trade_path = str(xch_names[0]) + "-->" + str(xch_names[1])
                    # buy at xch1 and sell at xch0
                    else:
                        trade_path = str(xch_names[1]) + "-->" + str(xch_names[0])

                    if abs(over_dict[i]["gross_perc_profit"][-1]) > abs(
                        best_set["trade_value"]
                    ):
                        # check market depth
                        best_set["trade_value"] = over_dict[i]["gross_perc_profit"][-1]
                        best_set["pair_name"] = i
                        best_set["trade_path"] = trade_path
                        best_set["base_token"] = base_token_address
                        best_set["other_token"] = other_token_address
                        best_set["xch0"] = config[blockchain][xch_names[0]][
                            "pool_pairs"
                        ][i]
                        best_set["xch1"] = config[blockchain][xch_names[1]][
                            "pool_pairs"
                        ][i]

                else:
                    # do not trade
                    trade_path = ""

                over_dict[i]["trade_path"].append(trade_path)

            else:
                if skip_pair[i] <= hour:
                    skip_pair[i] += 1
                else:
                    skip_pair[i] = 0

    # if the export file already exists then update it - otherwise create an export file
    file_path = "./Outputs/scanned_pairs_results.xlsx"

    if os.path.exists(file_path) == False:
        writer = pd.ExcelWriter(file_path)
        for i in over_dict:
            df = pd.DataFrame(over_dict[i])
            df.to_excel(writer, sheet_name=i)
            writer.save()
        writer.close()
        book = load_workbook(file_path)
        book.save(file_path)

    else:
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(
            file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        )
        for i in over_dict:
            df = pd.DataFrame(over_dict[i])
            df.to_excel(writer, sheet_name=i)
            book.save(file_path)

    print("")
    print("##########################################")
    print("")
    print("   Download Complete!")
    print("")
    print("##########################################")
    print("")


def scan_by_ID(
    primary_dex, secondary_dex, blockchain, selected_ids, save_name, base_token
):
    print("")
    col_list = [
        "DEX_pool_no",
        "t0_address",
        "t1_address",
        "DEX_name",
        "pool_address",
        "t0_reserves",
        "t1_reserves",
        "amountOut",
        "s_DEX_name",
        "s_pool_address",
        "s_t0_reserves",
        "s_t1_reserves",
        "s_amountOut",
        "balance",
        "arb",
    ]
    return_list = []

    wb = Workbook()
    ws = wb.active
    ws.append(col_list)
    wb.save(save_name)

    # load factory abi json
    with open("./ABIs/" + str(primary_dex) + "_factory.json", "r") as file:
        factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(primary_dex) + "_factory_pool.json", "r") as file:
        pool_abi = json.loads(file.read())

    with open("./ABIs/" + str(primary_dex) + "_router.json", "r") as file:
        router_abi = json.loads(file.read())

    # load factory abi json
    with open("./ABIs/" + str(secondary_dex) + "_factory.json", "r") as file:
        sdex_factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(secondary_dex) + "_factory_pool.json", "r") as file:
        sdex_pool_abi = json.loads(file.read())

    with open(
        "./ABIs/" + str(secondary_dex.split("_")[0]) + "_router.json", "r"
    ) as file:
        sdex_router_abi = json.loads(file.read())

    config = load_config()
    w3 = Web3(Web3.HTTPProvider(config[blockchain]["network"]["mainnet"]))
    factory_address = w3.toChecksumAddress(
        config[blockchain][primary_dex][str(primary_dex) + "_factory"]
    )
    factory_contract = w3.eth.contract(abi=factory_abi, address=factory_address)

    dex_router_contract = w3.eth.contract(
        abi=router_abi,
        address=config[blockchain][primary_dex][str(primary_dex) + "_router"],
    )

    if selected_ids == None:
        record_length = factory_contract.functions.allPairsLength().call()
        sample_range = list(range(record_length))
    else:
        sample_range = selected_ids

    sdex_address = w3.toChecksumAddress(
        config[blockchain][secondary_dex][str(secondary_dex) + "_factory"]
    )
    sdex_contract = w3.eth.contract(abi=sdex_factory_abi, address=sdex_address)

    s_dex_router_contract = w3.eth.contract(
        abi=sdex_router_abi,
        address=config[blockchain][secondary_dex][str(secondary_dex) + "_router"],
    )

    count = 0
    SEARCHING = True
    while SEARCHING == True:
        nap = 300
        if count > 0:
            time.sleep(nap)

        for i in tqdm(sample_range, "Downloading: ", leave=False):
            pool_address = factory_contract.functions.allPairs(i).call()
            pool_address = w3.toChecksumAddress(pool_address)
            pool_contract = w3.eth.contract(abi=pool_abi, address=pool_address)

            # get pool data
            reserves = pool_contract.functions.getReserves().call()
            token0_address = pool_contract.functions.token0().call()
            token1_address = pool_contract.functions.token1().call()

            reserves[0] = reserves[0]
            reserves[1] = reserves[1]

            # see if the same pool exists in pancake swap - secondary_dex
            try:
                sdex_pool_address = sdex_contract.functions.getPair(
                    token0_address, token1_address
                ).call()
                sdex_pool_address = w3.toChecksumAddress(sdex_pool_address)
                sdex_pool_contract = w3.eth.contract(
                    abi=sdex_pool_abi, address=sdex_pool_address
                )

                # get secondary pool data
                s_reserves = sdex_pool_contract.functions.getReserves().call()
                s_reserves[0] = s_reserves[0]
                s_reserves[1] = s_reserves[1]

                # determine how many other tokens the base token will get
                base_token_in = 1
                base_token_in = Web3.toWei(base_token_in, "ether")
                for addy in [token0_address, token1_address]:
                    if addy != base_token:
                        other_token = addy
                        break
                trade_path = [base_token, other_token]
                amountOut = dex_router_contract.functions.getAmountsOut(
                    base_token_in, trade_path
                ).call()[1]
                s_amountOut = s_dex_router_contract.functions.getAmountsOut(
                    base_token_in, trade_path
                ).call()[1]

                # find the better value
                if amountOut > s_amountOut:
                    end_trade = s_dex_router_contract.functions.getAmountsOut(
                        amountOut, [other_token, base_token]
                    ).call()[1]
                elif amountOut < s_amountOut:
                    end_trade = dex_router_contract.functions.getAmountsOut(
                        s_amountOut, [other_token, base_token]
                    ).call()[1]
                else:
                    end_trade = s_dex_router_contract.functions.getAmountsOut(
                        amountOut, [other_token, base_token]
                    ).call()[1]

                arb = (end_trade - base_token_in) / base_token_in
                arb = arb - 0.00166

                # if other_token == "0xacFC95585D80Ab62f67A14C566C1b7a49Fe91167":
                #     arb = arb - 0.02
                # else:
                #     arb = arb - 0.00166

                if arb > 0:
                    return_list = [
                        i,
                        token0_address,
                        token1_address,
                        primary_dex,
                        pool_address,
                        reserves[0],
                        reserves[1],
                        amountOut,
                        secondary_dex,
                        sdex_pool_address,
                        s_reserves[0],
                        s_reserves[1],
                        s_amountOut,
                        end_trade,
                        arb,
                    ]

                    book = load_workbook(save_name)
                    sheet = book.active
                    sheet.append(return_list)
                    book.save(save_name)

                    for i in range(20):
                        play_obj = simpleaudio.WaveObject.from_wave_file(
                            "mixkit-basketball-buzzer-1647.wav"
                        ).play()
                        play_obj.wait_done()
                        time.sleep(1)

                    SEARCHING = False

            except:
                pass

        count += 1

        print(f"Cycle {count} complete")

    print("")
    print("##########################################")
    print("")
    print("   Download Complete!")
    print("")
    print("##########################################")
    print("")


def blind_scan(
    primary_dex,
    secondary_dex,
    blockchain,
    save_name,
    base_token,
    small_cap_threshold,
    exchange,
):
    print("")
    col_list = [
        "DEX_pool_no",
        "t0_address",
        "t1_address",
        "DEX_name",
        "pool_address",
        "pool_size",
        "amountOut",
        "s_DEX_name",
        "s_pool_address",
        "s_pool_size",
        "s_amountOut",
        "balance",
        "arb",
    ]
    return_list = []

    wb = Workbook()
    ws = wb.active
    ws.append(col_list)
    wb.save(save_name)

    # load factory abi json
    with open("./ABIs/" + str(primary_dex) + "_factory.json", "r") as file:
        factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(primary_dex) + "_factory_pool.json", "r") as file:
        pool_abi = json.loads(file.read())

    with open("./ABIs/" + str(primary_dex) + "_router.json", "r") as file:
        router_abi = json.loads(file.read())

    # load factory abi json
    with open("./ABIs/" + str(secondary_dex) + "_factory.json", "r") as file:
        sdex_factory_abi = json.loads(file.read())

    # load pool sample abi json
    with open("./ABIs/" + str(secondary_dex) + "_factory_pool.json", "r") as file:
        sdex_pool_abi = json.loads(file.read())

    with open(
        "./ABIs/" + str(secondary_dex.split("_")[0]) + "_router.json", "r"
    ) as file:
        sdex_router_abi = json.loads(file.read())

    config = load_config()
    w3 = Web3(Web3.HTTPProvider(config[blockchain]["network"]["mainnet"]))
    factory_address = w3.toChecksumAddress(
        config[blockchain][primary_dex][str(primary_dex) + "_factory"]
    )
    factory_contract = w3.eth.contract(abi=factory_abi, address=factory_address)

    dex_router_contract = w3.eth.contract(
        abi=router_abi,
        address=config[blockchain][primary_dex][str(primary_dex) + "_router"],
    )

    record_length = factory_contract.functions.allPairsLength().call()
    sample_range = list(range(record_length))

    sdex_address = w3.toChecksumAddress(
        config[blockchain][secondary_dex][str(secondary_dex) + "_factory"]
    )
    sdex_contract = w3.eth.contract(abi=sdex_factory_abi, address=sdex_address)

    s_dex_router_contract = w3.eth.contract(
        abi=sdex_router_abi,
        address=config[blockchain][secondary_dex][str(secondary_dex) + "_router"],
    )

    for i in tqdm(sample_range, "Downloading: ", leave=False):
        # for i in sample_range:
        # init(autoreset=True)
        # step_no = i + 1
        # # step_perc = step_no / len(sample_range)
        # # step_perc = step_perc * 100
        # # step_perc = round(step_perc, 2)
        # print(Fore.YELLOW + f"--> {step_no} of {len(sample_range)} completed <--")

        # for i in sample_range:
        pool_address = factory_contract.functions.allPairs(i).call()
        pool_address = w3.toChecksumAddress(pool_address)
        pool_contract = w3.eth.contract(abi=pool_abi, address=pool_address)

        # get pool data
        token0_address = pool_contract.functions.token0().call()
        token1_address = pool_contract.functions.token1().call()

        # check if either of the tokens is a base token
        # if it isn't then skip to the next pool
        if base_token in [token0_address, token1_address]:
            # print("Match!")
            # see if the same pool exists in both DEXes
            try:
                sdex_pool_address = sdex_contract.functions.getPair(
                    token0_address, token1_address
                ).call()
                sdex_pool_address = w3.toChecksumAddress(sdex_pool_address)
                sdex_pool_contract = w3.eth.contract(
                    abi=sdex_pool_abi, address=sdex_pool_address
                )

                # get primary pool data
                reserves = pool_contract.functions.getReserves().call()

                # get secondary pool data
                s_reserves = sdex_pool_contract.functions.getReserves().call()

                # which is base and which is other
                base_token_in = 1
                base_token_in = Web3.toWei(base_token_in, "ether")
                token_count = 0
                for addy in [token0_address, token1_address]:
                    if addy != base_token:
                        other_token = addy
                    else:
                        pool_value = reserves[token_count] / (10**18)
                        s_pool_value = s_reserves[token_count] / (10**18)
                    token_count += 1

                # skip known bad other tokens
                # bad_other_tokens = ["0xacFC95585D80Ab62f67A14C566C1b7a49Fe91167"]
                bad_other_tokens = []
                if other_token not in bad_other_tokens:
                    # check that both pools are above the threshold
                    cond1 = small_cap_threshold == None
                    cond2 = (
                        pool_value > small_cap_threshold
                        and s_pool_value > small_cap_threshold
                    )

                    if cond1 or cond2:
                        # determine how many other tokens the base token will get
                        trade_path = [base_token, other_token]
                        amountOut = dex_router_contract.functions.getAmountsOut(
                            base_token_in, trade_path
                        ).call()[1]
                        s_amountOut = s_dex_router_contract.functions.getAmountsOut(
                            base_token_in, trade_path
                        ).call()[1]

                        # find the better value
                        if amountOut > s_amountOut:
                            end_trade = s_dex_router_contract.functions.getAmountsOut(
                                amountOut, [other_token, base_token]
                            ).call()[1]
                        elif amountOut < s_amountOut:
                            end_trade = dex_router_contract.functions.getAmountsOut(
                                s_amountOut, [other_token, base_token]
                            ).call()[1]
                        else:
                            end_trade = s_dex_router_contract.functions.getAmountsOut(
                                amountOut, [other_token, base_token]
                            ).call()[1]

                        profit_loss = end_trade - base_token_in
                        pl_perc = (profit_loss / base_token_in) * 100
                        # if profit_loss < 0:
                        #     init(autoreset=True)
                        #     print(
                        #         Fore.RED
                        #         + f"Do Not Trade -- LOSS at {round(pl_perc, 2)}%"
                        #     )
                        # else:
                        #     pass
                        # print(f"Starting = {base_token_in} and Ending = {end_trade}")

                        arb = ((amountOut - s_amountOut) / amountOut) * 100
                        abs_arb = abs(arb)

                        # if other_token == "0xacFC95585D80Ab62f67A14C566C1b7a49Fe91167":
                        #     arb = arb - 0.02
                        # else:
                        #     arb = arb - 0.00166

                        # only record those with arbitrage value
                        if abs_arb != 0 and pl_perc > 0:

                            if arb >= 0:
                                exchange_path = [exchange[0], exchange[1]]
                            else:
                                exchange_path = [exchange[1], exchange[0]]

                            print("")
                            init(autoreset=True)
                            print(Fore.GREEN + "##############################")
                            print(
                                Fore.GREEN
                                + f"Trade found at a PROFIT of {round(pl_perc, 2)}% with an Arbitrage value of {round(arb, 2)}%"
                            )
                            print(
                                Fore.GREEN + f"Trading {base_token} for {other_token}"
                            )
                            print(
                                Fore.GREEN
                                + f"Minimum pool size is {small_cap_threshold} --- Primary DEX pool size is {round(pool_value, 0)} and Secondary DEX pool size is {round(s_pool_value, 0)}"
                            )
                            print(
                                Fore.GREEN
                                + f"Get {amountOut} from Primary and {s_amountOut} from Secondary DEX pools"
                            )
                            print(
                                Fore.GREEN
                                + f"Buy from {exchange_path[0]} and sell to {exchange_path[1]}"
                            )
                            print(Fore.GREEN + "##############################")
                            print("")

                            return_list = [
                                i,
                                token0_address,
                                token1_address,
                                primary_dex,
                                pool_address,
                                pool_value,
                                amountOut,
                                secondary_dex,
                                sdex_pool_address,
                                s_pool_value,
                                s_amountOut,
                                end_trade,
                                arb,
                            ]

                            # book = load_workbook(save_name)
                            # sheet = book.active
                            # sheet.append(return_list)
                            # book.save(save_name)

            except:
                pass

    init(autoreset=True)
    print("")
    print("##########################################")
    print("")
    print("   Download Complete!")
    print("")
    print("##########################################")
    print("")
